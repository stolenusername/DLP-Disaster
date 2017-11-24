from openpyxl import Workbook
import os
from openpyxl.styles import Font

directory = "./DLP-Disaster"
os.mkdir( directory, 755 );

data = [ { "name": "Kevin Flynn", "address": "123 east first", "city": "Houston", "state": "Texas", "employer": "Evil Corp", "pay": "$100,000", "card": "4111 1111 1111 1111", "cvv": 548, "expiration": "03/25/2030","account": "79097377399", "routing": "011103093", "social": "481-60-6432", "medical": "fractured tibia"},
{ "name": "Rick Deckard", "address": "4585 west 3rd", "city": "Cameron", "state": "Texas", "employer": "Umbrella Corporation", "pay": "$110,000", "card": "5500 0000 0000 0004", "cvv": 541, "expiration": "05/15/2018","account": "58988383370", "routing": "067014822", "social": "132-38-6608", "medical": "influenza"},
{ "name": "Eldon Tyrell", "address": "78789 north star drive", "city": "Rochester", "state": "New York", "employer": "Evil Corp", "pay": "$120,000", "card": "3400 0000 0000 009", "cvv": 742, "expiration": "10/31/2020","account": "77865418918", "routing": "211274450", "social": "648-44-9876", "medical": "tinitus"},
{ "name": "David Banner", "address": "123 east first", "city": "Las Vegas", "state": "Nevada", "employer": "Tyrell Coropration", "pay": "$130,000", "card": "6011 0000 0000 0004", "cvv": 133, "expiration": "06/01/2019","account": "34471510845", "routing": "211370545", "social": "542-98-0130", "medical": "laceration"},
{ "name": "Lisbeth Salander", "address": "45646 dead end street", "city": "Fresno", "state": "California", "employer": "Encom Corporation", "pay": "$140,000", "card": "2223000010476510", "cvv": 812, "expiration": "07/04/2021","account": "62490496395", "routing": "054001725", "social": "626-71-7961", "medical": "concussion"},
{ "name": "Sarah Connor", "address": "48554 pecan st", "city": "North Shore", "state": "Texas", "employer": "Cyberdyne Systems", "pay": "$150,000", "card": "371449635398431", "cvv": 732, "expiration": "12/21/2024","account": "19781088865", "routing": "011400071", "social": "643-32-6499", "medical": "torn Anterior cruciate ligament"},
{ "name": "Abby Sciuto", "address": "525 NW 138th st", "city": "San Antonio", "state": "Texas", "employer": "Stark Industries", "pay": "$160,000", "card": "4061724061724061", "cvv": 574, "expiration": "01/01/2025","account": "55983292078", "routing": "031201360", "social": "043-44-2390", "medical": "torn medial collateral ligament"},
{ "name": "Tony Stark", "address": "566 5th ave", "city": "Edmond", "state": "Oklahoma", "employer": "Big Kahuna Burger", "pay": "$170,000", "card": "4061624061424061", "cvv": 548, "expiration": "08/08/2030","account": "55683252071", "routing": "042000013", "social": "431-54-6432", "medical": "broken ribs" }
]



wb = Workbook()
ws = wb.active
ws_sheet = wb.get_sheet_by_name('Sheet')
ws_sheet.title = 'Users Credit Cards'

a = ws['A1']
a.font = Font(bold=True)
ws['A1'] = 'Name'
ws.column_dimensions["A"].width = 12

b = ws['B1']
b.font = Font(bold=True)
ws['B1'] = 'Number'
ws.column_dimensions["B"].width = 22

c = ws['C1']
c.font = Font(bold=True)
ws['C1'] = 'CVV'
ws.column_dimensions["C"].width = 30

d = ws['D1']
d.font = Font(bold=True)
ws['D1'] = 'Expiration'
ws.column_dimensions["D"].width = 10




for i in range(len(data)):
	user = (data[i]['name'])
	ccard = data[i]['card']
	ccardcvv = data[i]['cvv']
	expiry = data[i]['expiration']
	ws.append([user, ccard, ccardcvv, expiry])
	wb.save("./DLP-Disaster/CreditCards.xlsx")


    





